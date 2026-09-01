"""Loads the Birbal reference lender dataset (100 fictional-but-realistic banks,
one row per bank, matching birbal.club/find-best-lender's real question flow)
into the database, replacing whatever is currently loaded.

Source file shape (see the "README" and "Coverage Stats" sheets in the workbook
itself for the full column reference and matching rules this data was designed
against): one row per bank, with Employment_Types_Accepted, Documents_Accepted,
and Property_Types_Accepted as pipe-separated lists, and one income-threshold
column per employment type (Min_Net_Monthly_Salary_INR, Min_Annual_Turnover_INR,
Min_Annual_Gross_Receipts_INR, Min_Monthly_Pension_INR) since a bank checks a
different income measure depending on which employment type it's evaluating.

Because the right income check depends on employment type, a bank that accepts
N employment types becomes N separate HomeLoanProductModel rows here, one per
employment type it serves, each carrying only the one income rule that applies —
"a bank's criteria genuinely differ by employment type" is modeled as separate
products, not a scoping column on one product's rules (see app/database.py's
EligibilityRuleModel docstring for why).

Run with: python -m app.load_birbal_dataset <path-to-xlsx>
Clears existing banks/products/attributes/rules first — this is meant to fully
replace whatever's currently loaded, not add to it.
"""

import asyncio
import re
import sys

import openpyxl
from sqlalchemy import delete, select

from app.database import (
    AttributeModel,
    BankModel,
    EligibilityRuleModel,
    HomeLoanProductModel,
    async_session_factory,
    create_all_tables,
)

# Display strings in the source file -> the internal values the rest of the app
# uses (see app/schemas.py's EmploymentType/PropertyType/DocumentType enums).
EMPLOYMENT_TYPE_MAP = {
    "Salaried": "salaried",
    "Self-employed": "self_employed",
    "Professional": "professional",
    "Pensioner": "pensioner",
}
PROPERTY_TYPE_MAP = {
    "Standard urban property": "standard_urban",
    "Semi-urban village": "semi_urban_village",
    "Under construction": "under_construction",
    "Others": "others",
}
DOCUMENT_MAP = {
    "ITR/Form16": "itr_form16",
    "ITR": "itr",
    "GST": "gst",
    "Business proof": "business_proof",
    "Salary slip": "salary_slip",
    "Cash income": "cash_income",
    "Pension proof": "pension_proof",
}

# Which income attribute + source column applies for each employment type — the
# one piece of "which rule to use depends on another answer" logic this loader
# encodes, matching the README's own matching rules exactly.
INCOME_BY_EMPLOYMENT_TYPE = {
    "salaried": ("net_monthly_salary", "Min_Net_Monthly_Salary_INR"),
    "self_employed": ("annual_turnover", "Min_Annual_Turnover_INR"),
    "professional": ("annual_gross_receipts", "Min_Annual_Gross_Receipts_INR"),
    "pensioner": ("monthly_pension", "Min_Monthly_Pension_INR"),
}

ATTRIBUTE_CATALOG = [
    {"key": "cibil_score", "label": "CIBIL score", "category": "Credit", "data_type": "number"},
    {"key": "employment_type", "label": "Employment type", "category": "Income", "data_type": "text"},
    {"key": "net_monthly_salary", "label": "Net monthly salary", "category": "Income", "data_type": "number"},
    {"key": "annual_turnover", "label": "Annual business turnover", "category": "Income", "data_type": "number"},
    {
        "key": "annual_gross_receipts",
        "label": "Annual gross receipts",
        "category": "Income",
        "data_type": "number",
    },
    {"key": "monthly_pension", "label": "Monthly pension", "category": "Income", "data_type": "number"},
    {"key": "has_co_borrower", "label": "Has co-borrower", "category": "Income", "data_type": "boolean"},
    {"key": "property_type", "label": "Property type", "category": "Property", "data_type": "text"},
    # Not a real borrower answer — see domain.py's "any_of" operator.
    {"key": "documents_any_of", "label": "Any income document", "category": "Documents", "data_type": "text"},
    {"key": "min_loan_amount", "label": "Minimum loan amount", "category": "Loan Amount", "data_type": "number"},
    {"key": "max_loan_amount", "label": "Maximum loan amount", "category": "Loan Amount", "data_type": "number"},
    {"key": "interest_rate_pct", "label": "Interest rate (from)", "category": "Pricing", "data_type": "number"},
    {"key": "interest_rate_range", "label": "Interest rate range", "category": "Pricing", "data_type": "text"},
    {"key": "processing_fee", "label": "Processing fee", "category": "Pricing", "data_type": "text"},
    {"key": "lender_type", "label": "Lender type", "category": "Pricing", "data_type": "text"},
]


def _split(value: str | None, mapping: dict[str, str]) -> list[str]:
    if not value:
        return []
    return [mapping[part.strip()] for part in str(value).split("|") if part.strip()]


def _parse_rate_floor(rate_range: str) -> str:
    """'8.71%-10.2%' -> '8.71' (the lower bound, used for scoring/display —
    see domain.py's WeightedScoringStrategy, which expects a single number)."""
    match = re.match(r"\s*([\d.]+)%", str(rate_range))
    if not match:
        raise ValueError(f"Could not parse a rate floor out of {rate_range!r}")
    return match.group(1)


def read_banks(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Banks"]
    headers = [cell.value for cell in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    banks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx["Bank_Name"]]:
            continue
        banks.append({h: row[i] for h, i in idx.items()})
    return banks


async def load(xlsx_path: str) -> None:
    banks = read_banks(xlsx_path)
    await create_all_tables()

    async with async_session_factory() as session:
        # Only wipe rows this loader itself created (source="excel_import") —
        # banks added by other loaders (e.g. app/seed_calibration_banks.py,
        # tagged "manual_calibration") are a different `source` and must
        # survive this reload untouched. See BankModel's docstring for why.
        excel_bank_ids = (
            (await session.execute(select(BankModel.id).where(BankModel.source == "excel_import")))
            .scalars()
            .all()
        )
        if excel_bank_ids:
            excel_product_ids = (
                (
                    await session.execute(
                        select(HomeLoanProductModel.id).where(HomeLoanProductModel.bank_id.in_(excel_bank_ids))
                    )
                )
                .scalars()
                .all()
            )
            if excel_product_ids:
                await session.execute(
                    delete(EligibilityRuleModel).where(EligibilityRuleModel.product_id.in_(excel_product_ids))
                )
                await session.execute(
                    delete(HomeLoanProductModel).where(HomeLoanProductModel.id.in_(excel_product_ids))
                )
            await session.execute(delete(BankModel).where(BankModel.id.in_(excel_bank_ids)))
        await session.flush()

        # The attribute catalog is shared with other loaders' rules (e.g. the
        # calibration banks' eligibility_rules point at these same attribute
        # rows) — so this only adds attributes that don't exist yet, matched
        # by key, and never deletes or recreates ones that already exist.
        attributes = {a.key: a for a in (await session.execute(select(AttributeModel))).scalars().all()}
        for attr in ATTRIBUTE_CATALOG:
            if attr["key"] not in attributes:
                model = AttributeModel(
                    key=attr["key"], label=attr["label"], category=attr["category"], data_type=attr["data_type"]
                )
                session.add(model)
                attributes[attr["key"]] = model
        await session.flush()

        product_count = 0
        rule_count = 0
        for entry in banks:
            bank = BankModel(name=entry["Bank_Name"], source="excel_import")
            session.add(bank)
            await session.flush()

            employment_types = _split(entry["Employment_Types_Accepted"], EMPLOYMENT_TYPE_MAP)
            property_types = _split(entry["Property_Types_Accepted"], PROPERTY_TYPE_MAP)
            documents = _split(entry["Documents_Accepted"], DOCUMENT_MAP)
            rate_floor = _parse_rate_floor(entry["Interest_Rate_Range"])

            for emp_type in employment_types:
                income_attr, income_col = INCOME_BY_EMPLOYMENT_TYPE[emp_type]
                income_threshold = entry[income_col]
                if income_threshold is None:
                    # Listed as accepted but no threshold given for it — data
                    # inconsistency in the source row; skip rather than guess.
                    continue

                label = {v: k for k, v in EMPLOYMENT_TYPE_MAP.items()}[emp_type]
                product = HomeLoanProductModel(bank_id=bank.id, product_name=f"Home Loan — {label}")
                session.add(product)
                await session.flush()
                product_count += 1

                rules = [
                    EligibilityRuleModel(
                        product_id=product.id,
                        attribute_id=attributes["cibil_score"].id,
                        operator="between",
                        value=f"{entry['Min_CIBIL']},{entry['Max_CIBIL']}",
                    ),
                    EligibilityRuleModel(
                        product_id=product.id,
                        attribute_id=attributes["employment_type"].id,
                        operator="==",
                        value=emp_type,
                    ),
                    EligibilityRuleModel(
                        product_id=product.id,
                        attribute_id=attributes[income_attr].id,
                        operator=">=",
                        value=str(income_threshold),
                    ),
                    EligibilityRuleModel(
                        product_id=product.id,
                        attribute_id=attributes["property_type"].id,
                        operator="in",
                        value=",".join(property_types),
                    ),
                    EligibilityRuleModel(
                        product_id=product.id,
                        attribute_id=attributes["documents_any_of"].id,
                        operator="any_of",
                        value=",".join(documents),
                    ),
                    # Facts — stored values, not conditions. See domain.py's get_fact.
                    EligibilityRuleModel(
                        product_id=product.id,
                        attribute_id=attributes["min_loan_amount"].id,
                        operator="fact",
                        value=str(entry["Min_Loan_Amount_INR"]),
                    ),
                    EligibilityRuleModel(
                        product_id=product.id,
                        attribute_id=attributes["max_loan_amount"].id,
                        operator="fact",
                        value=str(entry["Max_Loan_Amount_INR"]),
                    ),
                    EligibilityRuleModel(
                        product_id=product.id,
                        attribute_id=attributes["interest_rate_pct"].id,
                        operator="fact",
                        value=rate_floor,
                    ),
                    EligibilityRuleModel(
                        product_id=product.id,
                        attribute_id=attributes["interest_rate_range"].id,
                        operator="fact",
                        value=str(entry["Interest_Rate_Range"]),
                    ),
                    EligibilityRuleModel(
                        product_id=product.id,
                        attribute_id=attributes["processing_fee"].id,
                        operator="fact",
                        value=str(entry["Processing_Fee"]),
                    ),
                    EligibilityRuleModel(
                        product_id=product.id,
                        attribute_id=attributes["lender_type"].id,
                        operator="fact",
                        value=str(entry["Lender_Type"]),
                    ),
                ]

                # Co-borrower requirement only exists for Pensioner products, and
                # only "Yes" actually blocks a match — "No"/"Optional" never do
                # (see the source workbook's README, rule 5).
                if emp_type == "pensioner" and entry["CoBorrower_Required_For_Pensioner"] == "Yes":
                    rules.append(
                        EligibilityRuleModel(
                            product_id=product.id,
                            attribute_id=attributes["has_co_borrower"].id,
                            operator="required",
                            value="true",
                        )
                    )

                session.add_all(rules)
                rule_count += len(rules)

        await session.commit()
        print(f"Loaded {len(banks)} banks, {product_count} products, {rule_count} eligibility rules.")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python -m app.load_birbal_dataset <path-to-xlsx>")
        sys.exit(1)
    asyncio.run(load(sys.argv[1]))
