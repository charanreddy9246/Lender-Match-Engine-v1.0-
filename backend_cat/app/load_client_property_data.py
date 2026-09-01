"""Loads the client's property-eligibility workbook into the Lender_Matching
database on our own Postgres server — a completely different data source
from app/load_birbal_dataset.py (which loads our own fictional Birbal
dataset into Supabase), kept as its own file for exactly that reason: each
client's raw file looks nothing like another's, so each gets its own
translator script, even though both write into the same table design (see
backend/db_schemas/README.md).

Source file shape: one sheet per income/employment type (Salaried,
Self-Employed, Pensioner, Cash Income, NRI), same 23 real banks listed in
every sheet, each row a Yes/No checklist: does this bank serve this income
type, and for which property classification, sub-type, usage, stage, and
location.

This is only the property side of the client's data — no CIBIL score, loan
amount, income threshold, or interest rate yet (the client hasn't provided
that numeric data). Those get added later as "fact" rules on the same
product rows this script creates, once that data arrives — see
domain.py's WeightedScoringStrategy, which already scores a product with no
pricing facts as a neutral 0 rather than failing, so this half-loaded state
works fine in the meantime (only the bias table produces any real ranking
until then).

Connects directly to the Postgres server using the POSTGRES_HOST/PORT/USER/
PASSWORD/DB values in the root .env — NOT settings.database_url — so this
script always targets the right database regardless of whatever the main
app's DATABASE_URL happens to be pointed at that day.

Run with: python -m app.load_client_property_data <path-to-xlsx>
Safe to re-run — only wipes rows this script itself created
(source="client_property_import"), matching the same source-tagging
convention as app/database.py's BankModel.
"""

import asyncio
import os
import sys
import urllib.parse

import openpyxl
from dotenv import load_dotenv
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.database import (
    AttributeModel,
    Base,
    BankModel,
    EligibilityRuleModel,
    HomeLoanProductModel,
)

SOURCE = "client_property_import"

# Sheet name in the workbook -> our internal employment_type key + a human
# label for the product name. Cash Income and NRI don't exist in the old
# Birbal dataset's employment types — that's fine, this is just data, the
# matching engine doesn't hardcode which employment types exist.
SHEETS = {
    "Salaried": ("salaried", "Salaried"),
    "Self-Employed": ("self_employed", "Self-Employed"),
    "Pensioner": ("pensioner", "Pensioner"),
    "Cash Income": ("cash_income", "Cash Income"),
    "NRI": ("nri", "NRI"),
}

# Column number (1-indexed, matching row 4 of every sheet) -> internal
# property_type value. Column 3 (the income-type Yes/No) and columns
# 18-27 (usage/stage/location) are handled separately below.
PROPERTY_TYPE_COLUMNS = {
    4: "residential_vacant_land",
    5: "residential_apartment",
    6: "residential_independent_building",
    7: "residential_semi_independent_uds",
    8: "commercial_farm_land",
    9: "commercial_vacant_land",
    10: "commercial_independent_building",
    11: "commercial_semi_independent_uds",
    12: "commercial_temporary_structure",
    13: "industrial_vacant_land",
    14: "industrial_warehouse",
    15: "res_cum_comm_independent_building",
    16: "res_cum_comm_building",
    17: "res_cum_comm_multi_unit",
}
USAGE_COLUMNS = {18: "self_occupied", 19: "let_out", 20: "lease"}
STAGE_COLUMNS = {21: "new_purchase", 22: "resale", 23: "under_construction", 24: "take_over"}
LOCATION_COLUMNS = {25: "standard_urban", 26: "peri_urban", 27: "rural"}

ATTRIBUTE_CATALOG = [
    {"key": "employment_type", "label": "Employment type", "category": "Income", "data_type": "text"},
    {"key": "property_type", "label": "Property type", "category": "Property", "data_type": "text"},
    {"key": "property_usage", "label": "Property usage", "category": "Property", "data_type": "text"},
    {"key": "property_stage", "label": "Property stage", "category": "Property", "data_type": "text"},
    {"key": "property_location", "label": "Property location", "category": "Property", "data_type": "text"},
]

# Row 4 is the header row (SL No, Banks, <income type>, <27 Yes/No columns);
# real bank rows start at row 5. Some sheets have trailing blank/footnote
# rows after the last bank (e.g. Self-Employed's methodology note) — those
# have no bank name in column 2, so the loop below skips them naturally.
FIRST_DATA_ROW = 5


def _lender_matching_engine():
    """A separate engine from app/database.py's module-level `engine` on
    purpose — that one follows settings.database_url (whatever the main
    app's .env currently points at), and this script must always write to
    Lender_Matching specifically, regardless of that.
    """
    load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))
    host = os.environ["POSTGRES_HOST"]
    port = os.environ["POSTGRES_PORT"]
    user = os.environ["POSTGRES_USER"]
    password = urllib.parse.quote(os.environ["POSTGRES_PASSWORD"], safe="")
    db = os.environ["POSTGRES_DB"]
    url = f"postgresql+asyncpg://{user}:{password}@{host}:{port}/{db}"
    return create_async_engine(url, echo=False, connect_args={"statement_cache_size": 0})


def read_sheet(ws) -> list[dict]:
    rows = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        bank_name_raw = ws.cell(row=r, column=2).value
        if not bank_name_raw or not str(bank_name_raw).strip():
            continue
        rows.append(
            {
                "bank_name": str(bank_name_raw).strip(),
                "accepted": ws.cell(row=r, column=3).value == "Yes",
                "property_type": [
                    key for col, key in PROPERTY_TYPE_COLUMNS.items() if ws.cell(row=r, column=col).value == "Yes"
                ],
                "usage": [key for col, key in USAGE_COLUMNS.items() if ws.cell(row=r, column=col).value == "Yes"],
                "stage": [key for col, key in STAGE_COLUMNS.items() if ws.cell(row=r, column=col).value == "Yes"],
                "location": [
                    key for col, key in LOCATION_COLUMNS.items() if ws.cell(row=r, column=col).value == "Yes"
                ],
            }
        )
    return rows


async def load(xlsx_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Expected sheet {sheet_name!r} not found in workbook")

    engine = _lender_matching_engine()
    session_factory = async_sessionmaker(engine, expire_on_commit=False)

    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    async with session_factory() as session:
        # Only wipe rows this script itself created — see module docstring.
        existing_bank_ids = (
            (await session.execute(select(BankModel.id).where(BankModel.source == SOURCE))).scalars().all()
        )
        if existing_bank_ids:
            existing_product_ids = (
                (
                    await session.execute(
                        select(HomeLoanProductModel.id).where(HomeLoanProductModel.bank_id.in_(existing_bank_ids))
                    )
                )
                .scalars()
                .all()
            )
            if existing_product_ids:
                await session.execute(
                    delete(EligibilityRuleModel).where(EligibilityRuleModel.product_id.in_(existing_product_ids))
                )
                await session.execute(
                    delete(HomeLoanProductModel).where(HomeLoanProductModel.id.in_(existing_product_ids))
                )
            await session.execute(delete(BankModel).where(BankModel.id.in_(existing_bank_ids)))
        await session.flush()

        attributes = {a.key: a for a in (await session.execute(select(AttributeModel))).scalars().all()}
        for attr in ATTRIBUTE_CATALOG:
            if attr["key"] not in attributes:
                model = AttributeModel(
                    key=attr["key"], label=attr["label"], category=attr["category"], data_type=attr["data_type"]
                )
                session.add(model)
                attributes[attr["key"]] = model
        await session.flush()

        banks_by_name: dict[str, BankModel] = {}
        product_count = 0
        rule_count = 0

        for sheet_name, (emp_key, emp_label) in SHEETS.items():
            for entry in read_sheet(wb[sheet_name]):
                if not entry["accepted"]:
                    continue

                bank = banks_by_name.get(entry["bank_name"])
                if bank is None:
                    bank = BankModel(name=entry["bank_name"], source=SOURCE)
                    session.add(bank)
                    await session.flush()
                    banks_by_name[entry["bank_name"]] = bank

                product = HomeLoanProductModel(bank_id=bank.id, product_name=f"Home Loan — {emp_label}")
                session.add(product)
                await session.flush()
                product_count += 1

                rules = [
                    EligibilityRuleModel(
                        product_id=product.id, attribute_id=attributes["employment_type"].id, operator="==", value=emp_key
                    )
                ]
                for field, attr_key in (
                    ("property_type", "property_type"),
                    ("usage", "property_usage"),
                    ("stage", "property_stage"),
                    ("location", "property_location"),
                ):
                    values = entry[field]
                    if values:
                        rules.append(
                            EligibilityRuleModel(
                                product_id=product.id,
                                attribute_id=attributes[attr_key].id,
                                operator="in",
                                value=",".join(values),
                            )
                        )

                session.add_all(rules)
                rule_count += len(rules)

        await session.commit()
        print(
            f"Loaded {len(banks_by_name)} banks, {product_count} products, {rule_count} eligibility rules "
            f"into Lender_Matching. No pricing data yet — every product will score as neutral until "
            f"the client's numeric data arrives."
        )

    await engine.dispose()


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python -m app.load_client_property_data <path-to-xlsx>")
        sys.exit(1)
    asyncio.run(load(sys.argv[1]))
